import streamlit as st
import pandas as pd
import re
import io
import os
from openpyxl import load_workbook
from shutil import copyfile

st.set_page_config(page_title="G-Change｜Google企業リスト整形ツール", layout="centered")

# ヘッダー
st.title("📄 G-Change：Google企業リスト整形ツール")
st.caption("ファイルをアップロードするだけで、自動整形＋テンプレートに反映します。")

# テンプレートファイルパス（アプリ内に格納済み）
TEMPLATE_PATH = "template.xlsx"
TEMPLATE_SHEET = "入力マスター"

# ファイルアップロード（ドラッグ＆ドロップ対応）
uploaded_file = st.file_uploader("📤 整形したいExcelファイルをアップロードしてください", type=["xlsx"])

def to_half_width(s):
    return s.translate(str.maketrans("０１２３４５６７８９－ー―", "0123456789---"))

def is_structured_format(df):
    return list(df.columns[:4]) == ['企業名', '業種', '住所', '電話番号']

def clean_vertical_list(filepath):
    wb = load_workbook(filename=filepath, data_only=True)
    ws = wb.active
    rows = [cell[0].value for cell in ws.iter_rows(min_col=1, max_col=1)]

    companies = []
    for i in range(3, len(rows)):
        line = rows[i]
        if not isinstance(line, str):
            continue
        line_half = to_half_width(line)
        phone_match = re.search(r'\d{2,4}-\d{2,4}-\d{3,4}', line_half)
        if phone_match:
            phone = phone_match.group()
            address = rows[i - 1] if isinstance(rows[i - 1], str) else ''
            industry = rows[i - 2] if isinstance(rows[i - 2], str) else ''
            company = rows[i - 3] if isinstance(rows[i - 3], str) else ''
            industry = re.sub(r'^\d\.\d\(\d+\)\s*·\s*', '', industry).strip()
            companies.append({
                '企業名': str(company).strip(),
                '業種': str(industry).strip(),
                '住所': str(address).strip(),
                '電話番号': phone
            })

    return pd.DataFrame(companies)

def remove_duplicates(df):
    df_dedup = df[df['電話番号'].notna()]
    df_unique = df_dedup.drop_duplicates(subset='電話番号', keep='first')
    df_empty = df[df['電話番号'].isna()]
    final_df = pd.concat([df_unique, df_empty], ignore_index=True)
    return final_df

if uploaded_file:
    filename = uploaded_file.name
    base_name = os.path.splitext(filename)[0]
    extension = "：リスト.xlsx"

    with open("uploaded.xlsx", "wb") as f:
        f.write(uploaded_file.read())

    try:
        df = pd.read_excel("uploaded.xlsx")
        # 整形済み or 入力マスター付きのパターン
        if is_structured_format(df):
            st.info("✅ 整形済みファイルとして処理します（重複削除）")
            df = remove_duplicates(df)
        elif '入力マスター' in pd.ExcelFile("uploaded.xlsx").sheet_names:
            st.info("✅ テンプレートファイルとして処理します（入力マスターから抽出）")
            df = pd.read_excel("uploaded.xlsx", sheet_name='入力マスター', skiprows=1, usecols="B:E", names=['企業名', '業種', '住所', '電話番号'])
            df = df.dropna(subset=['企業名'])
            df = remove_duplicates(df)
        else:
            st.info("🔄 Google検索縦型リストとして処理中（並べ替え）")
            df = clean_vertical_list("uploaded.xlsx")
    except Exception as e:
        st.error(f"❌ 読み込みエラー: {e}")
        st.stop()

    # テンプレートに書き込む
    output_file = f"{base_name}：リスト.xlsx"
    copyfile(TEMPLATE_PATH, output_file)

    book = load_workbook(output_file)
    sheet = book[TEMPLATE_SHEET]

    for i, row in df.iterrows():
        sheet.cell(row=i+2, column=2, value=row['企業名'])
        sheet.cell(row=i+2, column=3, value=row['業種'])
        sheet.cell(row=i+2, column=4, value=row['住所'])
        sheet.cell(row=i+2, column=5, value=row['電話番号'])

    book.save(output_file)

    st.success(f"✅ 整形完了！{len(df)}件の企業データをテンプレートに反映しました。")
    with open(output_file, "rb") as f:
        st.download_button("📥 ダウンロード（テンプレート反映済みファイル）", f.read(), file_name=output_file)
